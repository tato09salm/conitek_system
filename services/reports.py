import pandas as pd
from fpdf import FPDF
import streamlit as st
from config import Config
import os
from io import BytesIO
from datetime import datetime
from PIL import Image, ImageDraw, ImageFont
import qrcode


# ─────────────────────────────────────────────────────────────────────────────
# Helpers compartidos por todos los constructores de PDF
# ─────────────────────────────────────────────────────────────────────────────

def _safe(text) -> str:
    """Convierte texto arbitrario a cadena segura para Latin-1 (codificación FPDF)."""
    if text is None:
        return ""
    s = str(text)
    replacements = {
        "\u2026": "...", "\u2014": "-", "\u2013": "-",
        "\u201c": '"',  "\u201d": '"',
        "\u2018": "'",  "\u2019": "'",
    }
    for k, v in replacements.items():
        s = s.replace(k, v)
    try:
        s.encode("latin-1")
        return s
    except UnicodeEncodeError:
        return s.encode("latin-1", "ignore").decode("latin-1")


class _BrandedPDF(FPDF):
    """PDF con encabezado y pie de página institucionales UNT/CONITEK."""

    # Paleta UNT
    NAVY   = (0,   0,   128)
    GOLD   = (255, 215,   0)
    WHITE  = (255, 255, 255)
    LGRAY  = (238, 244, 255)   # fila alterna
    DKGRAY = (80,  80,   80)
    ROWBDR = (200, 210, 228)   # borde suave entre filas

    def __init__(self, report_title: str, orientation: str = "L"):
        super().__init__(orientation=orientation, unit="mm", format="A4")
        self._report_title = report_title
        self._gen_date = datetime.now().strftime("%d/%m/%Y  %H:%M")
        self.set_auto_page_break(auto=True, margin=14)
        self.set_margins(10, 30, 10)

    def header(self):
        # Banda navy de fondo
        self.set_fill_color(*self.NAVY)
        self.rect(0, 0, self.w, 27, "F")
        # Línea dorada inferior de la banda
        self.set_draw_color(*self.GOLD)
        self.set_line_width(0.8)
        self.line(0, 26.5, self.w, 26.5)
        self.set_line_width(0.2)
        # Nombre del congreso (dorado, pequeño)
        self.set_xy(10, 2)
        self.set_font("Arial", "B", 7)
        self.set_text_color(*self.GOLD)
        self.cell(self.w - 20, 5, _safe(Config.CONGRESS_NAME), align="C")
        # Título del reporte (blanco, prominente)
        self.set_xy(10, 9)
        self.set_font("Arial", "B", 13)
        self.set_text_color(*self.WHITE)
        self.cell(self.w - 20, 8, _safe(self._report_title), align="C")
        # Fecha generación (gris claro, pequeño)
        self.set_xy(self.w - 65, 19)
        self.set_font("Arial", "", 6.5)
        self.set_text_color(190, 210, 255)
        self.cell(55, 5, _safe(f"Generado: {self._gen_date}"), align="R")
        # Restaurar defaults
        self.set_text_color(0, 0, 0)
        self.set_draw_color(0, 0, 0)

    def footer(self):
        self.set_y(-10)
        self.set_draw_color(*self.NAVY)
        self.set_line_width(0.35)
        self.line(10, self.get_y(), self.w - 10, self.get_y())
        self.ln(1)
        self.set_font("Arial", "", 7)
        self.set_text_color(*self.DKGRAY)
        self.cell((self.w - 20) / 2, 5, _safe(Config.LOCATION), align="L")
        self.cell(0, 5, _safe(f"Pagina  {self.page_no()}"), align="R")
        self.set_text_color(0, 0, 0)
        self.set_line_width(0.2)
        self.set_draw_color(0, 0, 0)


class ReportService:
    @staticmethod
    def generate_excel(data, filename, headers=None):
        df = pd.DataFrame(data)
        os.makedirs(Config.REPORTS_DIR, exist_ok=True)
        output_path = os.path.join(Config.REPORTS_DIR, filename)
        df.to_excel(output_path, index=False, sheet_name='Reporte')
        return output_path

    @staticmethod
    def generate_pdf(data, filename, title="Reporte"):
        pdf = FPDF(orientation='L', unit='mm', format='A4')
        pdf.set_auto_page_break(auto=True, margin=12)
        pdf.add_page()
        pdf.set_font("Arial", 'B', 16)
        pdf.cell(0, 10, title, ln=True, align='C')
        pdf.set_font("Arial", size=10)
        pdf.ln(6)
        if not data:
            os.makedirs(Config.REPORTS_DIR, exist_ok=True)
            output_path = os.path.join(Config.REPORTS_DIR, filename)
            pdf.output(output_path)
            return output_path
        headers = list(data[0].keys())
        # Cálculo de anchos por contenido
        padding = 4
        max_width = pdf.w - 20
        content_widths = []
        for h in headers:
            max_str = len(str(h))
            for row in data:
                max_str = max(max_str, len(str(row.get(h, ""))))
            w = min(max(pdf.get_string_width("W") * max_str + padding, 24), 95)
            content_widths.append(w)
        total = sum(content_widths)
        if total > max_width:
            scale = max_width / total
            content_widths = [w * scale for w in content_widths]
        # Encabezados
        pdf.set_fill_color(200, 220, 255)
        pdf.set_font("Arial", 'B', 10)
        x0 = pdf.get_x()
        for w, htxt in zip(content_widths, headers):
            pdf.cell(w, 8, str(htxt), border=1, fill=True, align='L')
        pdf.ln()
        # Filas con altura uniforme por fila
        pdf.set_font("Arial", size=10)
        line_h = 6
        def wrap_text(txt, width):
            s = str(txt)
            if " " not in s:
                return [s]
            words = s.split(" ")
            lines, cur = [], ""
            for w in words:
                test = (cur + " " + w).strip()
                if pdf.get_string_width(test) + padding <= width:
                    cur = test
                else:
                    if cur:
                        lines.append(cur)
                        cur = w
                    else:
                        lines.append(w)
                        cur = ""
            if cur:
                lines.append(cur)
            return lines or [""]
        for row in data:
            cell_lines = []
            for w, h in zip(content_widths, headers):
                lines = wrap_text(row.get(h, ""), w)
                cell_lines.append(lines)
            max_lines = max(len(lines) for lines in cell_lines)
            row_h = line_h * max_lines
            x = x0
            y = pdf.get_y()
            for w, lines in zip(content_widths, cell_lines):
                if len(lines) < max_lines:
                    lines = lines + [""] * (max_lines - len(lines))
                pdf.set_xy(x, y)
                pdf.multi_cell(w, line_h, "\n".join(lines), border=1, align='L')
                x += w
            pdf.set_y(y + row_h)
        os.makedirs(Config.REPORTS_DIR, exist_ok=True)
        output_path = os.path.join(Config.REPORTS_DIR, filename)
        pdf.output(output_path)
        return output_path
    
    @staticmethod
    def df_to_excel_bytes(df: pd.DataFrame):
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Reporte')
            try:
                ws = writer.sheets['Reporte']
                from openpyxl.utils import get_column_letter
                for i, col in enumerate(df.columns, 1):
                    max_len = max(
                        [len(str(col))] + [len(str(v)) for v in df[col].astype(str).tolist()]
                    )
                    ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 50)
            except Exception:
                pass
        return output.getvalue()
    
    @staticmethod
    def data_to_pdf_bytes(data, title="Reporte"):
        """Genera PDF con diseño institucional UNT/CONITEK (tabla con encabezados navy/dorado,
        filas alternas, pie de página con número de hoja)."""

        def _wrap(txt, cell_w, pdf_obj):
            s = _safe(txt)
            words = s.split()
            if not words:
                return [""]
            lines, cur = [], ""
            for word in words:
                test = (cur + " " + word).strip()
                if pdf_obj.get_string_width(test) + 5 <= cell_w:
                    cur = test
                else:
                    if cur:
                        lines.append(cur)
                    cur = word
            if cur:
                lines.append(cur)
            return lines or [""]

        def _draw_col_headers(pdf_obj, col_widths, headers, x0):
            pdf_obj.set_fill_color(*_BrandedPDF.NAVY)
            pdf_obj.set_text_color(*_BrandedPDF.WHITE)
            pdf_obj.set_font("Arial", "B", 8)
            pdf_obj.set_xy(x0, pdf_obj.get_y())
            for cw, h in zip(col_widths, headers):
                pdf_obj.cell(cw, 8, _safe(str(h)), border=0, fill=True, align="C")
            pdf_obj.ln()
            # Línea dorada bajo los encabezados
            pdf_obj.set_draw_color(*_BrandedPDF.GOLD)
            pdf_obj.set_line_width(0.7)
            pdf_obj.line(x0, pdf_obj.get_y(), x0 + sum(col_widths), pdf_obj.get_y())
            pdf_obj.set_line_width(0.2)
            pdf_obj.set_draw_color(*_BrandedPDF.ROWBDR)
            pdf_obj.set_text_color(0, 0, 0)

        pdf = _BrandedPDF(report_title=title, orientation="L")
        pdf.add_page()

        if not data:
            pdf.set_font("Arial", "I", 10)
            pdf.set_text_color(80, 80, 80)
            pdf.cell(0, 10, "Sin datos para mostrar.", ln=True, align="C")
            return bytes(pdf.output(dest="S"))

        headers = list(data[0].keys())
        x0 = pdf.l_margin
        max_width = pdf.w - pdf.l_margin - pdf.r_margin

        # ── Calcular anchos de columna ──────────────────────────────────────
        pdf.set_font("Arial", "", 8)
        col_widths = []
        for h in headers:
            max_chars = min(len(str(h)), 35)
            for row in data:
                max_chars = max(max_chars, min(len(str(row.get(h, ""))), 35))
            w = max(pdf.get_string_width("W") * max_chars + 6, 18)
            col_widths.append(w)
        total = sum(col_widths)
        if total > max_width:
            scale = max_width / total
            col_widths = [w * scale for w in col_widths]

        # ── Encabezados de columna ──────────────────────────────────────────
        _draw_col_headers(pdf, col_widths, headers, x0)

        # ── Filas de datos ──────────────────────────────────────────────────
        line_h = 5
        pdf.set_font("Arial", "", 8)

        for row_idx, row in enumerate(data):
            cell_lines = [_wrap(row.get(h, ""), cw, pdf)
                          for cw, h in zip(col_widths, headers)]
            max_lines = max(len(l) for l in cell_lines)
            row_h = line_h * max_lines

            # Salto de página anticipado con re-impresión de cabeceras
            if pdf.get_y() + row_h > pdf.h - 14:
                pdf.add_page()
                _draw_col_headers(pdf, col_widths, headers, x0)
                pdf.set_font("Arial", "", 8)

            x, y = x0, pdf.get_y()

            # Fondo alterno
            fill_color = _BrandedPDF.WHITE if row_idx % 2 == 0 else _BrandedPDF.LGRAY
            pdf.set_fill_color(*fill_color)
            pdf.rect(x, y, sum(col_widths), row_h, "F")

            # Texto de cada celda
            pdf.set_text_color(15, 23, 42)
            for cw, lines in zip(col_widths, cell_lines):
                padded = lines + [""] * (max_lines - len(lines))
                pdf.set_xy(x + 1, y)
                pdf.multi_cell(cw - 1, line_h, "\n".join(padded),
                               border=0, fill=False, align="L")
                x += cw

            # Línea inferior suave
            pdf.set_draw_color(*_BrandedPDF.ROWBDR)
            pdf.line(x0, y + row_h, x0 + sum(col_widths), y + row_h)
            pdf.set_y(y + row_h)

        # Total de registros
        pdf.ln(2)
        pdf.set_font("Arial", "I", 7.5)
        pdf.set_text_color(*_BrandedPDF.DKGRAY)
        pdf.cell(0, 5, _safe(f"Total de registros: {len(data)}"), align="R")
        pdf.set_text_color(0, 0, 0)

        return bytes(pdf.output(dest="S"))
    
    @staticmethod
    def dashboard_pdf_bytes(figures, title="Dashboard CONITEK", kpis: dict | None = None):
        """PDF de dashboard con diseño institucional: portada con KPIs en tarjetas y
        una gráfica por página, todo con encabezado y pie de página UNT/CONITEK."""
        os.makedirs(Config.REPORTS_DIR, exist_ok=True)

        pdf = _BrandedPDF(report_title=title, orientation="P")
        pdf.add_page()

        # ── Tarjetas de KPIs ────────────────────────────────────────────────
        if kpis:
            pdf.set_font("Arial", "B", 11)
            pdf.set_text_color(*_BrandedPDF.NAVY)
            pdf.cell(0, 8, "Indicadores Clave de Rendimiento", ln=True, align="C")
            pdf.ln(3)

            kpi_items = list(kpis.items())
            page_w = pdf.w - pdf.l_margin - pdf.r_margin
            box_w = (page_w - 6) / 2
            box_h = 22

            for i, (k, v) in enumerate(kpi_items):
                col = i % 2
                if col == 0:
                    row_y = pdf.get_y()
                x = pdf.l_margin + col * (box_w + 6)

                # Fondo suave
                pdf.set_fill_color(238, 244, 255)
                pdf.rect(x, row_y, box_w, box_h, "F")
                # Acento navy izquierdo
                pdf.set_fill_color(*_BrandedPDF.NAVY)
                pdf.rect(x, row_y, 3.5, box_h, "F")
                # Valor (grande, navy)
                pdf.set_xy(x + 5.5, row_y + 2)
                pdf.set_font("Arial", "B", 17)
                pdf.set_text_color(*_BrandedPDF.NAVY)
                pdf.cell(box_w - 7, 12, _safe(str(v)), align="L")
                # Etiqueta (pequeña, gris)
                pdf.set_xy(x + 5.5, row_y + 14)
                pdf.set_font("Arial", "", 8)
                pdf.set_text_color(*_BrandedPDF.DKGRAY)
                pdf.cell(box_w - 7, 6, _safe(str(k)), align="L")

                if col == 1:
                    pdf.set_y(row_y + box_h + 5)

            # Si número impar de KPIs, avanzar
            if len(kpi_items) % 2 == 1:
                pdf.set_y(row_y + box_h + 5)

            pdf.set_text_color(0, 0, 0)
            pdf.ln(4)

        # ── Gráficas (una por página) ────────────────────────────────────────
        tmp_img_paths = []
        for idx, fig in enumerate(figures or []):
            try:
                img_bytes = fig.to_image(format="png", scale=2)
                img_path = os.path.join(Config.REPORTS_DIR, f"dash_fig_{idx+1}.png")
                with open(img_path, "wb") as f:
                    f.write(img_bytes)
                tmp_img_paths.append(img_path)
            except Exception:
                continue

        for img_path in tmp_img_paths:
            pdf.add_page()
            page_width = pdf.w - pdf.l_margin - pdf.r_margin
            pdf.image(img_path, x=pdf.l_margin, y=32, w=page_width)

        return bytes(pdf.output(dest="S"))
    
    @staticmethod
    def event_attendees_pdf_bytes(event_info: dict, speakers: list[str], audience: list[str]):
        pdf = FPDF(orientation='P', unit='mm', format='A4')
        pdf.set_auto_page_break(auto=True, margin=12)
        pdf.add_page()
        pdf.set_font("Arial", 'B', 16)
        pdf.cell(0, 10, "Lista de Participantes", ln=True, align='C')
        pdf.set_font("Arial", size=11)
        pdf.ln(2)
        # Normalizador a Latin-1 (evita errores por “…” y similares)
        def _ascii(s: str) -> str:
            if s is None:
                return ""
            txt = str(s)
            repl = {
                "…": "...",
                "—": "-",
                "–": "-",
                "“": '"',
                "”": '"',
                "‘": "'",
                "’": "'",
            }
            for k, v in repl.items():
                txt = txt.replace(k, v)
            try:
                txt.encode('latin-1')
                return txt
            except UnicodeEncodeError:
                return txt.encode('latin-1', 'ignore').decode('latin-1')
        pdf.cell(0, 7, _ascii(f"Evento: {event_info.get('name','-')}"), ln=True)
        pdf.cell(0, 7, _ascii(f"Lugar: {event_info.get('location','-')}    Fecha: {event_info.get('date','-')}"), ln=True)
        pdf.ln(3)
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(0, 8, "Ponentes", ln=True)
        pdf.set_font("Arial", 'B', 10)
        cw1, cw2, cw3 = 25, 60, 80
        pdf.set_fill_color(200, 220, 255)
        pdf.cell(cw1, 8, _ascii("DNI"), border=1, fill=True)
        pdf.cell(cw2, 8, _ascii("Ponente"), border=1, fill=True)
        pdf.cell(cw3, 8, _ascii("Nombre Ponencia"), border=1, ln=True, fill=True)
        pdf.set_font("Arial", size=10)
        if speakers:
            for row in speakers:
                parts = row.split(" ", 1)
                dni = parts[0] if len(parts) > 0 else ""
                rest = parts[1] if len(parts) > 1 else ""
                cols = rest.split(" - ", 1)
                ponente = cols[0] if len(cols) > 0 else ""
                titulo = cols[1] if len(cols) > 1 else ""
                y = pdf.get_y()
                x = pdf.get_x()
                pdf.multi_cell(cw1, 6, _ascii(dni), border=1)
                h = pdf.get_y() - y
                pdf.set_xy(x + cw1, y)
                pdf.multi_cell(cw2, 6, _ascii(ponente), border=1)
                h2 = pdf.get_y() - y
                pdf.set_xy(x + cw1 + cw2, y)
                pdf.multi_cell(cw3, 6, _ascii(titulo), border=1)
                hh = max(h, h2, pdf.get_y() - y)
                pdf.set_y(y + hh)
        else:
            pdf.cell(cw1 + cw2 + cw3, 8, _ascii("(sin registros)"), border=1, ln=True)
        pdf.ln(3)
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(0, 8, _ascii("Público"), ln=True)
        pdf.set_font("Arial", 'B', 10)
        pdf.cell(40, 8, _ascii("DNI"), border=1, fill=True)
        pdf.cell(125, 8, _ascii("Nombre"), border=1, ln=True, fill=True)
        pdf.set_font("Arial", size=10)
        if audience:
            for row in audience:
                parts = row.split(" ", 1)
                dni = parts[0] if len(parts) > 0 else ""
                nombre = parts[1] if len(parts) > 1 else ""
                y = pdf.get_y()
                x = pdf.get_x()
                pdf.multi_cell(40, 6, _ascii(dni), border=1)
                h = pdf.get_y() - y
                pdf.set_xy(x + 40, y)
                pdf.multi_cell(125, 6, _ascii(nombre), border=1)
                hh = max(h, pdf.get_y() - y)
                pdf.set_y(y + hh)
        else:
            pdf.cell(40 + 125, 8, _ascii("(sin registros)"), border=1, ln=True)
        return bytes(pdf.output(dest='S'))
    
    @staticmethod
    def payment_receipt_pdf_bytes(info: dict):
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", 'B', 18)
        pdf.cell(0, 10, "Comprobante de Pago", ln=True, align='C')
        pdf.set_font("Arial", size=11)
        pdf.ln(5)
        pdf.cell(0, 8, f"Evento: {info.get('event_name','-')}", ln=True)
        pdf.cell(0, 8, f"Referencia: {info.get('reference','-')}", ln=True)
        pdf.cell(0, 8, f"Comprobante N°: {info.get('receipt_no','-')}", ln=True)
        pdf.cell(0, 8, f"Fecha: {info.get('created_at','-')}", ln=True)
        pdf.ln(5)
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(0, 8, "Datos del Participante", ln=True)
        pdf.set_font("Arial", size=11)
        pdf.cell(0, 8, f"Nombre: {info.get('participant_name','-')}", ln=True)
        pdf.cell(0, 8, f"DNI: {info.get('dni','-')}", ln=True)
        pdf.ln(5)
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(0, 8, "Detalle del Pago", ln=True)
        pdf.set_font("Arial", size=11)
        pdf.cell(0, 8, f"Método: {info.get('method','-')}", ln=True)
        pdf.cell(0, 8, f"Monto: S/. {float(info.get('amount',0)):.2f}", ln=True)
        pdf.cell(0, 8, f"Estado: {info.get('status','-')}", ln=True)
        pdf.ln(10)
        pdf.cell(0, 8, f"Generado por: {Config.CONGRESS_NAME}", ln=True)
        return bytes(pdf.output(dest='S'))

    @staticmethod
    def dashboard_report_pdf_bytes(figures, captions: list[str] | None = None, title="Reporte del Dashboard", kpis: dict | None = None):
        os.makedirs(Config.REPORTS_DIR, exist_ok=True)
        pdf = FPDF(orientation='P', unit='mm', format='A4')
        pdf.set_auto_page_break(auto=False)
        pdf.add_page()
        pdf.set_font("Arial", 'B', 22)
        pdf.cell(0, 14, title, ln=True, align='C')
        pdf.set_font("Arial", size=12)
        pdf.cell(0, 8, f"{Config.CONGRESS_NAME}", ln=True, align='C')
        pdf.ln(6)
        pdf.set_font("Arial", size=11)
        pdf.cell(0, 8, f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}", ln=True, align='C')
        pdf.ln(10)
        if kpis:
            pdf.set_font("Arial", 'B', 14)
            pdf.cell(0, 10, "Indicadores Clave", ln=True)
            pdf.set_font("Arial", size=11)
            for k, v in kpis.items():
                pdf.cell(0, 8, f"- {k}: {v}", ln=True)
        try:
            import plotly.io as pio
        except Exception:
            return bytes(pdf.output(dest='S'))
        tmp_imgs = []
        for idx, fig in enumerate(figures or []):
            try:
                img_bytes = fig.to_image(format="png", scale=2)
                path = os.path.join(Config.REPORTS_DIR, f"dash_report_{idx+1}.png")
                with open(path, "wb") as f:
                    f.write(img_bytes)
                tmp_imgs.append(path)
            except Exception:
                continue
        page_num = 1
        for i, path in enumerate(tmp_imgs):
            pdf.add_page()
            pdf.set_font("Arial", 'B', 12)
            if captions and i < len(captions):
                pdf.cell(0, 8, captions[i], ln=True)
            else:
                pdf.cell(0, 8, "Gráfico", ln=True)
            pdf.image(path, x=10, y=20, w=190)
            pdf.set_xy(10, 285)
            pdf.set_font("Arial", size=9)
            pdf.cell(0, 6, f"{Config.CONGRESS_NAME} - Pagina {i+2}", ln=True, align='R')
            page_num += 1
        return bytes(pdf.output(dest='S'))

    @staticmethod
    def payment_voucher_pdf_bytes(info: dict):
        w, h = 1200, 900
        img = Image.new('RGB', (w, h), color='white')
        dr = ImageDraw.Draw(img)
        verde = (45, 122, 62)
        rojo = (255, 68, 68)
        negro = (0, 0, 0)
        gris = (51, 51, 51)
        try:
            ft = ImageFont.truetype("arial.ttf", 36)
            fs = ImageFont.truetype("arial.ttf", 30)
            ftx = ImageFont.truetype("arial.ttf", 26)
            frc = ImageFont.truetype("cour.ttf", 28)
            fdat = ImageFont.truetype("cour.ttf", 26)
            fserien = ImageFont.truetype("arialbd.ttf", 24)
            fserie = ImageFont.truetype("arial.ttf", 48)
        except:
            ft = fs = ftx = frc = fdat = fserien = fserie = ImageFont.load_default()
        dr.text((600, 60), "UNIVERSIDAD NACIONAL DE TRUJILLO", fill=verde, font=ft, anchor="mm")
        dr.text((600, 90), "DIRECCIÓN DE TESORERÍA", fill=verde, font=fs, anchor="mm")
        dr.text((600, 115), "ÁREA DE GESTIÓN DE INGRESOS", fill=verde, font=fs, anchor="mm")
        dr.text((600, 140), "Diego de Almagro N° 344 - TRUJILLO - PERÚ", fill=verde, font=ftx, anchor="mm")
        dr.text((600, 165), "R.U.C. 20172857628", fill=verde, font=ftx, anchor="mm")
        dr.text((600, 210), "RECIBO DE CAJA", fill=verde, font=ft, anchor="mm")
        # Mostrar la referencia del pago como número principal bajo el título
        ref_text = str(info.get("reference", "")) or str(info.get("receipt_no", ""))
        dr.text((650, 260), ref_text, fill=verde, font=ft, anchor="rm")
        y = 320
        dr.text((80, y), "TASA : 20  VENT:4      CTA.IP:122 .03.01 .01.01    SIAF:132.311", fill=negro, font=fdat)
        y += 40
        created = info.get("created_at","")
        fecha = created.split(" ")[0] if " " in created else created
        hora = created.split(" ")[1] if " " in created else ""
        monto = f"S/. {float(info.get('amount',0)):.2f}"
        dr.text((80, y), f"FECHA : {fecha}    HORA : {hora}    {monto}", fill=negro, font=fdat)
        y += 60
        nombre = info.get("participant_name","-")
        dni = info.get("dni","-")
        dr.text((80, y), f"HE RECIBIDO DE : {nombre} DNI: {dni}", fill=negro, font=fdat)
        y += 40
        evento = info.get("event_name","-")
        dr.text((80, y), f"EVENTO         : {evento}", fill=negro, font=fdat)
        y += 40
        dr.text((80, y), f"   LA SUMA DE   : {monto}", fill=negro, font=fdat)
        y += 40
        dr.text((80, y), "POR CONCEPTO DE:", fill=negro, font=fdat)
        y += 40
        concepto = f"Pago por {info.get('method','-')} - {evento}"
        dr.text((80, y), f"      {consepto if (consepto:=concepto) else concepto}", fill=negro, font=fdat)
        # ===== QR DE VERIFICACIÓN =====
        # Datos embebidos directamente en el QR (funciona sin internet/servidor)
        qr_content = (
            f"CONITEK 2026 - PAGO VERIFICADO\n"
            f"Recibo : {str(info.get('reference', '') or info.get('receipt_no', ''))}\n"
            f"Participante: {info.get('participant_name', '-')}\n"
            f"DNI  : {info.get('dni', '-')}\n"
            f"Monto: S/. {float(info.get('amount', 0)):.2f}\n"
            f"Metodo: {info.get('method', '-')}\n"
            f"Fecha: {fecha}"
        )
        qr = qrcode.QRCode(
            version=2,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=5,
            border=2,
        )
        qr.add_data(qr_content)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGB")
        qr_img = qr_img.resize((155, 155), Image.LANCZOS)
        qr_x, qr_y = 930, 645
        img.paste(qr_img, (qr_x, qr_y))
        dr.text((qr_x + 77, qr_y - 18), "VERIFICAR PAGO",
                fill=verde, font=fserien, anchor="mm")
        # ===== SERIE =====
        dr.text((80, 820), "Serie N° 1", fill=verde, font=fserien)
        # Mostrar REG-[ID] en la parte inferior derecha
        rec_id = str(info.get("receipt_no", "")).strip()
        reg_text = f"REG-{rec_id}" if rec_id else "REG-"
        dr.text((1100, 820), reg_text, fill=rojo, font=fserie, anchor="rm")
        for i in range(8):
            x = 1150
            yy = 100 + i * 100
            dr.ellipse([x-15, yy-15, x+15, yy+15], fill=gris)
        os.makedirs(Config.REPORTS_DIR, exist_ok=True)
        tmp = os.path.join(Config.REPORTS_DIR, f"voucher_{datetime.now().strftime('%Y%m%d%H%M%S%f')}.png")
        img.save(tmp, "PNG")
        # Ajustar el tamaño del PDF al aspecto real de la imagen (1200x900)
        # para evitar distorsión del QR y del resto del contenido
        pdf_w, pdf_h = 210, round(210 * h / w, 2)  # 210 x 157.5 mm
        pdf = FPDF(format=(pdf_w, pdf_h))
        pdf.set_auto_page_break(auto=False)
        pdf.add_page()
        pdf.image(tmp, x=0, y=0, w=pdf_w, h=pdf_h)
        return bytes(pdf.output(dest='S'))

def download_button_excel(df, filename, label="Descargar Excel"):
    processed_data = ReportService.df_to_excel_bytes(df)
    return st.download_button(
        label=label,
        data=processed_data,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
